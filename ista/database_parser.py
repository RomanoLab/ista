import csv
import os

import ipdb
import MySQLdb
import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm

from . import owl2
from .util import *


class DatabaseParser:
    """
    Base class for parsing the contents of a single source database and populating
    its contents into an OWL ontology.

    This class should not be instantiated directly; instead, you should instantiate
    one of its descendant classes that corresponds to the format of the data source.
    """

    def __init__(self, name, onto: owl2.Ontology):
        self.name = name
        self.ont = onto

    def parse_node_type(self, node_type: str, source_node_label: str, **kwargs):
        raise NotImplementedError("Error: Base classes should override this method.")

    def _merge_node(
        self,
        fields: dict,
        node_type: str,
        source_node_label: str,
        node_properties: dict,
        merge_column: dict,
        existing_class: str = None,
        skip_create_new_node: bool = False,
    ):
        if not existing_class:
            cl = get_onto_class_by_node_type(self.ont, node_type)
        else:
            cl = get_onto_class_by_node_type(self.ont, existing_class)
        if cl == None:
            raise RuntimeError(
                "Class label {0} not found in ComptoxAI ontology".format(node_type)
            )

        # Find node if it exists
        merge_col = merge_column["source_column_name"]
        merge_prop = merge_column["data_property"]

        # Search for existing individual with this property value
        merge_value = owl2.Literal(fields[merge_col])
        match = self.ont.search_by_data_property(merge_prop, merge_value)

        try:
            assert len(match) <= 1
        except AssertionError:
            # TODO: See Github issue 9
            match = [match[0]]

        # Only create a new individual if we didn't find a match
        if len(match) == 0:
            if skip_create_new_node:
                return
            individual_name = fields[source_node_label]
            # Create IRI for new individual
            individual_iri_str = safe_make_individual_name(individual_name, cl)
            onto_iri = self.ont.get_ontology_iri()
            if onto_iri:
                base_iri = onto_iri.value().get_namespace()
            else:
                base_iri = "http://example.org/onto#"
            individual_iri = owl2.IRI(base_iri + individual_iri_str)
            new_or_matched_individual = self.ont.create_individual(cl, individual_iri)
        else:
            new_or_matched_individual = match[0]

            # Make sure to append the new class label if we are supposed to
            if existing_class:
                new_class = get_onto_class_by_node_type(self.ont, node_type)
                self.ont.add_class_assertion(new_or_matched_individual, new_class)

        if node_properties:
            for field, d_prop in node_properties.items():
                # Make sure we don't overwrite the field we already matched on
                if field == merge_col:
                    continue
                value = (
                    owl2.Literal(fields[field]) if fields[field] is not None else None
                )
                safe_add_property(self.ont, new_or_matched_individual, d_prop, value)

    def _write_new_node(
        self,
        fields: dict,
        node_type: str,
        source_node_label: str,
        node_properties: dict = None,
    ):
        cl = get_onto_class_by_node_type(self.ont, node_type)

        individual_name = fields[source_node_label]

        # new_individual = cl(individual_name.lower())
        # Create IRI for new individual
        individual_iri_str = safe_make_individual_name(individual_name, cl)
        onto_iri = self.ont.get_ontology_iri()
        if onto_iri:
            base_iri = onto_iri.value().get_namespace()
        else:
            base_iri = "http://example.org/onto#"
        individual_iri = owl2.IRI(base_iri + individual_iri_str)
        new_individual = self.ont.create_individual(cl, individual_iri)

        if node_properties:
            for field, d_prop in node_properties.items():
                try:
                    value = (
                        owl2.Literal(fields[field])
                        if fields[field] is not None
                        else None
                    )

                    safe_add_property(self.ont, new_individual, d_prop, value)
                except KeyError:
                    # TODO: Log missing properties!
                    pass


class FlatFileDatabaseParser(DatabaseParser):
    """
    Parse a database saved to one or more flat files (e.g., CSV, TSV, etc.).

    Parameters
    ----------
    name : str
        Name of the database. The individual data files should be contained in a
        directory with the same name (case-sensitive).
    destination : owl2.Ontology
        Ontology to be populated with the database's contents.
    """

    def __init__(self, name: str, destination: owl2.Ontology, data_dir):
        super().__init__(name, destination)
        self.data_dir = data_dir

    def get_file_pointer_by_flatfile_name(self, filename):
        file_pointer = open(
            os.path.join(self.data_dir, self.name, filename),
            "r",
            encoding="utf-8",
            errors="ignore",
        )
        return file_pointer

    def get_file_pointer_by_node_label(self, node_label):
        raise NotImplementedError()

    def _store_table_headers(self, header_data, delimiter=","):
        headers = header_data.split(delimiter)

        print("FILE_HEADERS:")
        for h in headers:
            print(h)

        self.headers = headers

    def parse_relationship_type(
        self,
        relationship_type,
        source_filename: str,
        fmt: str,
        parse_config: dict,
        inverse_relationship_type=None,
        merge: bool = False,
        skip: bool = False,
    ):
        if skip:
            return

        no_rel_added = True

        print("PARSING RELATIONSHIP TYPE: {0}".format(relationship_type))
        print("FROM SOURCE DATABASE: {0}".format(self.name))

        s_node_type = parse_config["subject_node_type"]
        o_node_type = parse_config["object_node_type"]

        sub_match_prop_name = parse_config["subject_match_property"].name
        obj_match_prop_name = parse_config["object_match_property"].name

        if not fmt == "xlsx":
            fp = self.get_file_pointer_by_flatfile_name(source_filename)

        if fmt == "csv":
            reader = csv.reader(fp)

            if type(parse_config["headers"]) == list:
                headers = parse_config["headers"]
            elif parse_config["headers"] == True:
                headers = next(reader)

            if "skip_n_lines" in parse_config:
                for _ in range(parse_config["skip_n_lines"]):
                    next(reader)

            node_iter = reader

        elif fmt == "tsv":
            reader = csv.reader(fp, delimiter="\t")

            if type(parse_config["headers"]) == list:
                headers = parse_config["headers"]
            elif parse_config["headers"] == True:
                headers = next(reader)

            if "skip_n_lines" in parse_config:
                for _ in range(parse_config["skip_n_lines"]):
                    next(reader)

            node_iter = reader

        else:
            raise RuntimeError("Unsupported flatfile format: {0}".format(format))

        compound_fields = None
        if "compound_fields" in parse_config:
            compound_fields = list(parse_config["compound_fields"].keys())

        data_transforms = None
        if "data_transforms" in parse_config:
            data_transforms = list(parse_config["data_transforms"].keys())

        for rel in tqdm(node_iter):
            fields = dict(zip(headers, rel))

            if "filter_column" in parse_config:
                # Do a simple string matching. This isn't 'safe' and may lead
                # to issues down the line, but for now we are using it cautiously.
                # TODO: Refactor per above note
                if (
                    parse_config["filter_value"]
                    not in fields[parse_config["filter_column"]]
                ):
                    continue

            if compound_fields:
                for cf in compound_fields:
                    cf_config = parse_config["compound_fields"][cf]
                    cf_data = fields[cf].split(cf_config["delimiter"])
                    for subfield in cf_data:
                        s = subfield.split(cf_config["field_split_prefix"])
                        fields[s[0]] = s[-1]

            if data_transforms:
                for dt in data_transforms:
                    fields[dt] = parse_config["data_transforms"][dt](fields[dt])

            subj_ids = fields[parse_config["subject_column_name"]]
            obj_ids = fields[parse_config["object_column_name"]]

            if type(subj_ids) is not list:
                subj_ids = [subj_ids]
            if type(obj_ids) is not list:
                obj_ids = [obj_ids]

            for sid in subj_ids:
                for oid in obj_ids:
                    subject_match = self.ont.search_by_data_property(
                        parse_config["subject_match_property"], owl2.Literal(sid)
                    )
                    if len(subject_match) == 0:
                        continue
                    # try:
                    #     assert len(subject_match) == 1
                    # except AssertionError:
                    #     ipdb.set_trace()
                    #     print()

                    object_match = self.ont.search_by_data_property(
                        parse_config["object_match_property"], owl2.Literal(oid)
                    )
                    if len(object_match) == 0:
                        continue
                    # try:
                    #     assert len(object_match) == 1
                    # except AssertionError:
                    #     ipdb.set_trace()
                    #     print()

                    if no_rel_added:
                        no_rel_added = False

                    for sm in subject_match:
                        for om in object_match:
                            safe_add_property(self.ont, sm, relationship_type, om)
                            if inverse_relationship_type:
                                safe_add_property(
                                    self.ont, om, inverse_relationship_type, sm
                                )

        if no_rel_added:
            print("WARNING: NO RELATIONSHIPS ADDED TO ONTOLOGY")

    def parse_node_type(
        self,
        node_type: str,
        source_filename: str,
        fmt: str,
        parse_config: dict,
        merge: bool = True,
        append_class: bool = False,
        existing_class: str = None,
        skip_create_new_node: bool = False,
        skip: bool = False,
    ):
        """
        Parameters
        ----------
        node_type : str
          String name corresponding to the node type in the destination database.
        source_node_label : str
          String used to identify the node type in the source database.
        merge : bool
          When `True`, entities that already exist in the database will be *merged*
          with the new data. When `False`, a new entity will be created in the
          graph database regardless of whether a matching entity already exists in
          the database.
        """
        if skip:
            return

        no_node_added = True

        print("PARSING NODE TYPE: {0}".format(node_type))
        print("FROM SOURCE DATABASE: {0}".format(self.name))

        ont_class = get_onto_class_by_node_type(self.ont, node_type)

        if not fmt == "xlsx":
            fp = self.get_file_pointer_by_flatfile_name(source_filename)

        # Use the specified file format to create the following two items:
        # - `headers` - a list of column headers
        # - `node_iter` - an iterable that returns individual node records from the file
        #
        # Note: Each item returned by `node_iter` must be the same length as `headers`

        if fmt == "csv":
            reader = csv.reader(fp)

            if not parse_config["headers"]:
                raise NotImplementedError(
                    "Parsing of files without column headers is not currently supported"
                )
            if type(parse_config["headers"]) == list:
                headers = parse_config["headers"]
            elif parse_config["headers"] == True:
                headers = next(reader)

            if "skip_n_lines" in parse_config:
                for _ in range(parse_config["skip_n_lines"]):
                    next(reader)

            # for line in tqdm(reader):
            #     data = dict(zip(headers, line))
            node_iter = reader

        elif fmt == "tsv":
            reader = csv.reader(fp, delimiter="\t")

            if not parse_config["headers"]:
                raise NotImplementedError(
                    "Parsing of files without column headers is not currently supported"
                )
            headers = next(reader)

            # for line in tqdm(reader):
            #     data = dict(zip(headers, line))
            node_iter = reader

        elif fmt == "xlsx":
            sfn = os.path.join(self.data_dir, self.name, source_filename)
            print("LOADING .xlsx FILE - PLEASE BE PATIENT...")
            wb = load_workbook(filename=sfn, read_only=True, data_only=True)
            print("...done.")

            # Hopefully we only have one worksheet!
            ws = wb[wb.sheetnames[0]]

            header_cells = ws[1]
            headers = [hc.value for hc in header_cells]

            # for line in tqdm(ws.iter_rows(min_row=2)):
            #     vals = [l.value for l in line]
            #     data = dict(zip(headers, vals))
            node_iter = ws.iter_rows(min_row=2)

        elif fmt == "tsv-pandas":
            df = pd.read_csv(
                os.path.join(self.data_dir, self.name, source_filename), sep="\t"
            )

            headers = list(df.columns)

            # Use a generator expression to coerce the Pandas iterator into our desired format
            node_iter = (x for _, x in df.iterrows())

        else:
            raise RuntimeError("Unsupported flatfile format: {0}".format(format))

        compound_fields = None
        if "compound_fields" in parse_config:
            compound_fields = list(parse_config["compound_fields"].keys())

        data_transforms = None
        if "data_transforms" in parse_config:
            data_transforms = list(parse_config["data_transforms"].keys())

        if merge:
            merge_column = parse_config["merge_column"]
            if append_class:
                existing_class_label = existing_class
            else:
                existing_class_label = None

        for node in tqdm(node_iter):
            fields = dict(zip(headers, node))

            if "filter_column" in parse_config:
                # Do a simple string matching. This isn't 'safe' and may lead
                # to issues down the line, but for now we are using it cautiously.
                # TODO: Refactor per above note
                if (
                    parse_config["filter_value"]
                    not in fields[parse_config["filter_column"]]
                ):
                    continue

            if compound_fields:
                for cf in compound_fields:
                    cf_config = parse_config["compound_fields"][cf]
                    cf_data = fields[cf].split(cf_config["delimiter"])
                    for subfield in cf_data:
                        s = subfield.split(cf_config["field_split_prefix"])
                        fields[s[0]] = s[-1]

            if data_transforms:
                for dt in data_transforms:
                    fields[dt] = parse_config["data_transforms"][dt](fields[dt])

            if no_node_added:
                no_node_added = False

            if merge:
                self._merge_node(
                    fields,
                    node_type,
                    parse_config["iri_column_name"],
                    parse_config["data_property_map"],
                    merge_column,
                    existing_class_label,
                    skip_create_new_node,
                )
            else:
                self._write_new_node(
                    fields,
                    node_type,
                    parse_config["iri_column_name"],
                    parse_config["data_property_map"],
                )

        fp.close()

        if no_node_added:
            print("WARNING: NO NODES/PROPERTIES ADDED TO ONTOLOGY")


class MySQLDatabaseParser(DatabaseParser):
    """
    Parse a database stored on a MySQL server and populate its contents into
    an OWL ontology.

    The server must be locally accessible, and credentials should be passed at
    instantiation via the `config_dict` parameter (see below).

    Parameters
    ----------
    name : str
        Name of the database. Must be identical (case-sensitive) to the name
        of the database on the MySQL server.
    destination : owl2.Ontology
        Ontology to be populated with the database's contents.
    config_dict : dict
        Configuration details needed to connect to the MySQL database.
    """

    def __init__(self, name: str, destination: owl2.Ontology, config_dict: dict):
        super().__init__(name, destination)

        if "socket" in config_dict:
            self.conn = MySQLdb.Connection(
                host=config_dict["host"],
                user=config_dict["user"],
                password=config_dict["passwd"],
                database=name,
                unix_socket=config_dict["socket"],
            )
        else:
            self.conn = MySQLdb.Connection(
                host=config_dict["host"],
                user=config_dict["user"],
                password=config_dict["passwd"],
                database=name,
            )

    def parse_relationship_type(
        self,
        relationship_type,
        parse_config: dict,
        inverse_relationship_type=None,
        merge: bool = False,
        skip: bool = False,
    ):
        if skip:
            return

        no_rel_added = True

        print("PARSING RELATIONSHIP TYPE: {0}".format(relationship_type))
        print("FROM SOURCE DATABASE: {0}".format(self.name))

        s_node_type = parse_config["subject_node_type"]
        o_node_type = parse_config["object_node_type"]

        c = self.conn.cursor()

        if "custom_sql_query" in parse_config:
            fetch_lines_query = parse_config["custom_sql_query"]
        else:
            fetch_lines_query = "SELECT * FROM {0};".format(
                parse_config["source_table"]
            )
        c.execute(fetch_lines_query)

        headers = [cc[0] for cc in c.description]

        sub_match_prop_name = parse_config["subject_match_property"].name
        obj_match_prop_name = parse_config["object_match_property"].name

        for _ in tqdm(range(c.rowcount)):
            row = c.fetchone()

            fields = dict(zip(headers, row))

            if "compound_fields" in parse_config:
                for c_f_name, c_f_config in parse_config["compound_fields"].items():
                    fields[c_f_name] = fields[c_f_name].split(c_f_config["delimiter"])

            if "data_transforms" in parse_config:
                for dt_f_name, dt_f_transform in parse_config[
                    "data_transforms"
                ].items():
                    if type(fields[dt_f_name]) == list:
                        fields[dt_f_name] = [
                            dt_f_transform(x) for x in fields[dt_f_name]
                        ]
                    else:
                        fields[dt_f_name] = dt_f_transform(fields[dt_f_name])

            # Get subject and object nodes
            if parse_config["source_table_type"] == "foreignKey":
                subj_ids = fields[parse_config["subject_column_name"]]
                obj_ids = fields[parse_config["object_column_name"]]

                if type(subj_ids) is not list:
                    subj_ids = [subj_ids]
                if type(obj_ids) is not list:
                    obj_ids = [obj_ids]
            else:
                raise NotImplementedError("`source_table_type` not recognized")

            for sid in subj_ids:
                for oid in obj_ids:
                    subject_match = self.ont.search_by_data_property(
                        parse_config["subject_match_property"], owl2.Literal(sid)
                    )
                    if len(subject_match) == 0:
                        continue
                    try:
                        assert len(subject_match) == 1
                    except AssertionError:
                        ipdb.set_trace()
                        print()
                    object_match = self.ont.search_by_data_property(
                        parse_config["object_match_property"], owl2.Literal(oid)
                    )
                    if len(object_match) == 0:
                        continue
                    try:
                        assert len(object_match) == 1
                    except AssertionError:
                        ipdb.set_trace()
                        print()

                    if no_rel_added:
                        no_rel_added = False

                    safe_add_property(
                        subject_match[0], relationship_type, object_match[0]
                    )
                    if inverse_relationship_type:
                        safe_add_property(
                            object_match[0], inverse_relationship_type, subject_match[0]
                        )

        c.close()

        if no_rel_added:
            print("WARNING: NO RELATIONSHIPS ADDED TO ONTOLOGY")

    def parse_node_type(
        self,
        node_type: str,
        source_table: str,
        parse_config: dict,
        merge: bool = True,
        append_class: bool = False,
        existing_class: str = None,
        skip: bool = False,
    ):
        if skip:
            return

        no_node_added = True

        print("PARSING NODE TYPE: {0}".format(node_type))
        print("FROM SOURCE DATABASE: {0}".format(self.name))

        c = self.conn.cursor()

        if "custom_sql_query" in parse_config:
            fetch_lines_query = parse_config["custom_sql_query"]
        else:
            fetch_lines_query = "SELECT * FROM {0};".format(source_table)
        c.execute(fetch_lines_query)

        # Get column names
        headers = [cc[0] for cc in c.description]

        if merge:
            merge_column = parse_config["merge_column"]
            if append_class:
                existing_class_label = existing_class
            else:
                existing_class_label = None

        for _ in tqdm(range(c.rowcount)):  # would `while` be safer?
            row = c.fetchone()

            fields = dict(zip(headers, row))

            if "filter_column" in parse_config:
                # Do a simple string matching. This isn't 'safe' and may lead
                # to issues down the line, but for now we are using it cautiously.
                # TODO: Refactor per above note
                if (
                    parse_config["filter_value"]
                    not in fields[parse_config["filter_column"]]
                ):
                    continue

            if no_node_added:
                no_node_added = False

            if merge:
                self._merge_node(
                    fields=fields,
                    node_type=node_type,
                    source_node_label=parse_config["iri_column_name"],
                    node_properties=parse_config["data_property_map"],
                    merge_column=merge_column,
                    existing_class=existing_class_label,
                )
            else:
                self._write_new_node(
                    fields,
                    node_type,
                    parse_config["iri_column_name"],
                    parse_config["data_property_map"],
                )

        c.close()

        if no_node_added:
            print("WARNING: NO NODES/PROPERTIES ADDED TO ONTOLOGY")


# class XMLDatabaseParser(DatabaseParser):
